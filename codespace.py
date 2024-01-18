from openpyxl import load_workbook, Workbook
from openpyxl.chart import BarChart, Reference
from collections import Counter

# Load the original workbook
file_path = 'aptaujasdati.xlsx'  # replace with your file path
original_workbook = load_workbook(filename=file_path)
original_sheet = original_workbook.active

# Initialize a dictionary to hold the data
survey_data = {
    'ID': [],
    'AGE': [],
    'Answers': {f'Q{i+1}': [] for i in range(10)}
}

# Read the data from the original sheet
for row in original_sheet.iter_rows(min_row=2, values_only=True):
    survey_data['ID'].append(row[0])
    survey_data['AGE'].append(row[1])
    
    for i, answer in enumerate(row[2:], start=1):
        survey_data['Answers'][f'Q{i}'].append(answer)

# Calculate the most popular answer for each question
popular_answers = {}
for question, answers in survey_data['Answers'].items():
    answer_counter = Counter(answers)
    most_common_answer, count = answer_counter.most_common(1)[0]
    popular_answers[question] = {'Most Popular Answer': most_common_answer, 'Count': count}

# Create a new workbook to save results
new_workbook = Workbook()
new_sheet = new_workbook.active
new_sheet.title = "Original Data"

# Write the original data to the new sheet
for i, row in enumerate(original_sheet.iter_rows(values_only=True)):
    for j, cell in enumerate(row):
        new_sheet.cell(row=i+1, column=j+1, value=cell)

# Add a new sheet for analysis results
analysis_sheet = new_workbook.create_sheet(title="Analysis Results")

# Write headers
analysis_sheet.append(["Question", "Most Popular Answer", "Count"])

# Write the analysis results and create charts
for question, result in popular_answers.items():
    analysis_sheet.append([question, result['Most Popular Answer'], result['Count']])
    
    # Create a new sheet for the chart of this question
    chart_sheet = new_workbook.create_sheet(title=f"Chart {question}")
    
    # Count the frequency of each answer
    answer_counter = Counter(survey_data['Answers'][question])
    data = list(answer_counter.items())
    data.sort(key=lambda x: x[0])  # Sort by answer number
    
    # Write chart data to the chart sheet
    chart_sheet.append(["Answer", "Frequency"])
    for answer, frequency in data:
        chart_sheet.append([answer, frequency])
    
    # Create a bar chart
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = f"Answer Distribution for {question}"
    chart.x_axis.title = 'Answers'
    chart.y_axis.title = 'Frequency'
    
    # Reference the data for the chart
    categories = Reference(chart_sheet, min_col=1, min_row=2, max_row=len(data)+1)
    values = Reference(chart_sheet, min_col=2, min_row=1, max_row=len(data)+1)
    chart.add_data(values, titles_from_data=True)
    chart.set_categories(categories)
    
    # Add the chart to the sheet
    chart_sheet.add_chart(chart, "E2")

# Save the new workbook
new_workbook.save(filename='Result.xlsx')
